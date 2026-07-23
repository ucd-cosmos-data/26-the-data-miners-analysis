"""Issue the final go/no-go gate for downstream EDA.

This stage does not perform exploratory analysis or signal preprocessing. It
checks that the verified processed tree exposes readable, traceable inputs and
that every unresolved scientific warning remains visible to downstream users.
"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Sequence

from bci_core import (
    DEFAULT_CLEANED,
    DEFAULT_RAW,
    DEFAULT_REPORTS,
    archive_comparison_passed,
    assert_safe_paths,
    atomic_write_text,
    expected_participants,
    iter_files,
    read_csv_dicts,
)


CHECK_FIELDS = ["check", "status", "observed", "expected", "details"]
PARTICIPANT_PATTERN = re.compile(r"[ABC]\d+")


def _result(
    check: str,
    passed: bool,
    observed: object,
    expected: object,
    details: str,
) -> dict[str, str]:
    return {
        "check": check,
        "status": "pass" if passed else "fail",
        "observed": str(observed),
        "expected": str(expected),
        "details": details,
    }


def write_readiness_csv(
    path: Path,
    checks: Sequence[dict[str, str]],
) -> None:
    """Write the readiness checklist as stable UTF-8 with LF records."""

    output = io.StringIO(newline="")
    writer = csv.DictWriter(
        output,
        fieldnames=CHECK_FIELDS,
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    writer.writerows(checks)
    atomic_write_text(path, output.getvalue())


def workbook_participant_ids(path: Path) -> tuple[list[str], list[str]]:
    """Return participant IDs and sheet names without changing the workbook."""

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for the EDA-readiness gate") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    identifiers: list[str] = []
    try:
        for sheet in workbook.worksheets:
            for (value,) in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
                if isinstance(value, str) and PARTICIPANT_PATTERN.fullmatch(value.strip()):
                    identifiers.append(value.strip())
        return identifiers, list(workbook.sheetnames)
    finally:
        workbook.close()


def inspect_processed_csvs(
    cleaned: Path,
    manifest: Sequence[dict[str, str]],
) -> dict[str, Any]:
    """Check that every manifest-listed processed CSV is UTF-8, LF-only, and parseable."""

    manifest_by_cleaned = {
        row["cleaned_relative_path"]: row
        for row in manifest
        if row.get("cleaned_relative_path")
    }
    expected_paths = {
        relative for relative in manifest_by_cleaned if Path(relative).suffix.lower() == ".csv"
    }
    actual_paths = {
        path.relative_to(cleaned).as_posix()
        for path in iter_files(cleaned)
        if path.suffix.lower() == ".csv"
    }
    failures: list[str] = []
    parsed_rows = 0
    action_counts: Counter[str] = Counter()

    for relative in sorted(actual_paths):
        path = cleaned / relative
        row = manifest_by_cleaned.get(relative, {})
        action = row.get("action", "unlisted")
        action_counts[action] += 1
        payload = path.read_bytes()
        try:
            text = payload.decode("utf-8")
        except UnicodeDecodeError:
            failures.append(f"{relative}: not UTF-8")
            continue
        if b"\r" in payload:
            failures.append(f"{relative}: contains CR bytes")
        if action == "normalized_performance_csv":
            delimiter = ";"
        elif action == "normalized_frequency_csv":
            delimiter = ","
        else:
            try:
                delimiter = csv.Sniffer().sniff(text[:16384], delimiters=",;\t|").delimiter
            except csv.Error:
                delimiter = ","
        try:
            rows = list(csv.reader(io.StringIO(text, newline=""), delimiter=delimiter))
        except csv.Error as exc:
            failures.append(f"{relative}: CSV parse failed ({type(exc).__name__})")
            continue
        if not rows:
            failures.append(f"{relative}: no CSV records")
        parsed_rows += len(rows)

    missing = sorted(expected_paths - actual_paths)
    extras = sorted(actual_paths - expected_paths)
    failures.extend(f"{relative}: missing" for relative in missing)
    failures.extend(f"{relative}: not listed in cleaning manifest" for relative in extras)
    return {
        "expected_count": len(expected_paths),
        "actual_count": len(actual_paths),
        "parsed_rows": parsed_rows,
        "action_counts": dict(sorted(action_counts.items())),
        "failures": failures,
    }


def build_readiness_checks(
    raw: Path,
    cleaned: Path,
    reports: Path,
    rules_path: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, str]]:
    """Build the complete readiness checklist and return preserved warnings."""

    rules = json.loads(rules_path.read_text(encoding="utf-8"))
    inventory = read_csv_dicts(reports / "dataset_inventory.csv")
    manifest = read_csv_dicts(reports / "cleaning_manifest.csv")
    post_clean = read_csv_dicts(reports / "post_clean_validation.csv")
    issues = read_csv_dicts(reports / "validation_issues.csv")
    checks: list[dict[str, str]] = []

    archive_passed = archive_comparison_passed(reports / "archive_comparison.csv")
    checks.append(
        _result(
            "archive_verification",
            archive_passed,
            "passed" if archive_passed else "failed",
            "all Zenodo members match; only explicit administrative extras accepted",
            "Confirms the raw extraction used by the pipeline matches the published archive.",
        )
    )

    inventory_paths = {row.get("relative_path", "") for row in inventory}
    manifest_paths = {row.get("raw_relative_path", "") for row in manifest}
    manifest_verified = (
        bool(inventory)
        and len(manifest) == len(inventory)
        and manifest_paths == inventory_paths
        and all(row.get("status") == "verified" for row in manifest)
    )
    checks.append(
        _result(
            "cleaning_manifest",
            manifest_verified,
            f"{sum(row.get('status') == 'verified' for row in manifest)}/{len(manifest)} verified",
            f"{len(inventory)} unique raw files, all verified",
            "Requires exactly one verified lineage record for every inventoried raw file.",
        )
    )

    post_paths = {row.get("raw_relative_path", "") for row in post_clean}
    post_passed = (
        bool(post_clean)
        and len(post_clean) == len(inventory)
        and post_paths == inventory_paths
        and all(row.get("status") == "pass" for row in post_clean)
    )
    checks.append(
        _result(
            "post_clean_traceability",
            post_passed,
            f"{sum(row.get('status') == 'pass' for row in post_clean)}/{len(post_clean)} passed",
            f"{len(inventory)} raw immutability and cleaned-equivalence checks",
            "Covers raw SHA-256 identity, cleaned lineage, and GDF byte identity.",
        )
    )

    current_cleaned = {
        path.relative_to(cleaned).as_posix()
        for path in iter_files(cleaned)
        if path.relative_to(cleaned).as_posix() != ".gitkeep"
    }
    expected_cleaned = {
        row["cleaned_relative_path"]
        for row in manifest
        if row.get("cleaned_relative_path")
    }
    membership_ok = current_cleaned == expected_cleaned
    checks.append(
        _result(
            "processed_membership",
            membership_ok,
            f"{len(current_cleaned)} files; extras={len(current_cleaned - expected_cleaned)}; "
            f"missing={len(expected_cleaned - current_cleaned)}",
            f"exactly {len(expected_cleaned)} manifest-listed files",
            "Rejects untracked additions and missing processed files.",
        )
    )

    blocking = [row for row in issues if row.get("severity") in {"fatal", "error"}]
    warnings = [row for row in issues if row.get("severity") == "warning"]
    checks.append(
        _result(
            "blocking_validation_issues",
            not blocking,
            len(blocking),
            0,
            f"{len(warnings)} preserved warning(s) remain documented for EDA.",
        )
    )

    participant_dirs: set[str] = set()
    for group in ("DATA A", "DATA B", "DATA C"):
        group_path = cleaned / "Signals" / group
        if group_path.is_dir():
            participant_dirs.update(path.name for path in group_path.iterdir() if path.is_dir())
    expected_ids = expected_participants()
    checks.append(
        _result(
            "participant_directories",
            participant_dirs == expected_ids,
            f"{len(participant_dirs)} unique IDs",
            f"{len(expected_ids)} IDs: A1-A60, B61-B81, C82-C87",
            "Participant identifiers must remain unchanged between cleaning and EDA.",
        )
    )

    gdf_paths = {
        path.relative_to(cleaned).as_posix()
        for path in iter_files(cleaned)
        if path.suffix.lower() == ".gdf"
    }
    expected_gdf_count = int(rules["recordings"]["expected_gdf_files"])
    post_by_cleaned = {
        row.get("cleaned_relative_path", ""): row
        for row in post_clean
        if row.get("cleaned_relative_path")
    }
    gdf_traceable = all(
        path in post_by_cleaned
        and post_by_cleaned[path].get("status") == "pass"
        and post_by_cleaned[path].get("raw_current_sha256")
        == post_by_cleaned[path].get("cleaned_current_sha256")
        for path in gdf_paths
    )
    checks.append(
        _result(
            "gdf_handoff",
            len(gdf_paths) == expected_gdf_count and gdf_traceable,
            f"{len(gdf_paths)} files; byte-identical={gdf_traceable}",
            f"{expected_gdf_count} documented recordings, all traceable and byte-identical",
            "Signals remain unpreprocessed; EDA code should read them from data/processed/.",
        )
    )

    workbook_path = cleaned / "Perfomances.xlsx"
    workbook_error = ""
    workbook_ids: list[str] = []
    sheet_names: list[str] = []
    try:
        workbook_ids, sheet_names = workbook_participant_ids(workbook_path)
    except Exception as exc:
        workbook_error = f"{type(exc).__name__}: {exc}"
    workbook_ok = (
        not workbook_error
        and set(workbook_ids) == expected_ids
        and len(workbook_ids) == len(expected_ids)
    )
    checks.append(
        _result(
            "canonical_workbook",
            workbook_ok,
            workbook_error
            or f"{len(workbook_ids)} unique participant rows across {len(sheet_names)} sheet(s)",
            "readable workbook with exactly the 87 documented participant IDs",
            "Perfomances.xlsx is the canonical tabular source; its multi-section layout is preserved.",
        )
    )

    csv_result = inspect_processed_csvs(cleaned, manifest)
    csv_ok = (
        not csv_result["failures"]
        and csv_result["actual_count"] == csv_result["expected_count"]
    )
    checks.append(
        _result(
            "processed_csv_interfaces",
            csv_ok,
            f"{csv_result['actual_count']} files; {csv_result['parsed_rows']} parsed records; "
            f"failures={len(csv_result['failures'])}",
            f"{csv_result['expected_count']} manifest-listed UTF-8, LF-only, parseable CSV files",
            "; ".join(csv_result["failures"]) or "All normalized CSV interfaces are readable.",
        )
    )

    raw_resolved = raw.resolve(strict=True)
    cleaned_resolved = cleaned.resolve(strict=True)
    separated = (
        raw_resolved not in cleaned_resolved.parents
        and cleaned_resolved not in raw_resolved.parents
        and raw_resolved != cleaned_resolved
    )
    checks.append(
        _result(
            "raw_processed_separation",
            separated,
            f"raw={raw_resolved}; processed={cleaned_resolved}",
            "separate trees",
            "EDA must use data/processed/ and must never write into data/raw/.",
        )
    )
    documented_exceptions = {
        str(key): str(value)
        for key, value in rules.get("documented_exceptions", {}).items()
    }
    return checks, warnings, documented_exceptions


def write_readiness_report(
    destination: Path,
    cleaned: Path,
    checks: Sequence[dict[str, str]],
    warnings: Sequence[dict[str, str]],
    documented_exceptions: dict[str, str],
) -> str:
    verdict = "PASS" if all(row["status"] == "pass" for row in checks) else "FAIL"
    lines = [
        "# EDA Readiness",
        "",
        f"- Result: **{verdict}**",
        f"- Processed dataset: `{cleaned}`",
        f"- Checks passed: {sum(row['status'] == 'pass' for row in checks)}/{len(checks)}",
        f"- Blocking issues: {sum(row['status'] == 'fail' for row in checks)}",
        f"- Documented source exceptions: {len(documented_exceptions)}",
        f"- Open validation warnings: {len(warnings)}",
        "",
        "## Readiness checks",
        "",
        "| Check | Status | Observed |",
        "|---|---|---|",
    ]
    for row in checks:
        lines.append(
            f"| `{row['check']}` | **{row['status'].upper()}** | {row['observed']} |"
        )
    lines.extend(
        [
            "",
            "## EDA entry points",
            "",
            "- `data/processed/Perfomances.xlsx`: canonical participant, performance, questionnaire, and profile table. Preserve its three-section layout when importing.",
            "- `data/processed/Signals/DATA A|B|C/<participant>/*.gdf`: unpreprocessed EEG/EOG/EMG recordings. Read with MNE and keep participant/run identifiers.",
            "- `data/processed/Signals/**/frequency-band-selected*.csv`: UTF-8/LF-normalized MDFB outputs with scientific values unchanged.",
            "- `results/reports/cleaning_manifest.csv`: raw-to-processed lineage for every publisher file.",
            "",
            "## Documented source exceptions",
            "",
        ]
    )
    if documented_exceptions:
        for label, description in documented_exceptions.items():
            lines.append(f"- `{label}` — {description}")
    else:
        lines.append("- None.")
    lines.extend(["", "## Open validation warnings", ""])
    if warnings:
        for warning in warnings:
            label = warning.get("participant_run") or warning.get("relative_path") or "dataset"
            lines.append(
                f"- `{label}` — `{warning.get('check', '')}`: "
                f"{warning.get('observed_value', '')}. No value was changed."
            )
    else:
        lines.append("- None.")
    lines.extend(
        [
            "",
            "A `PASS` authorizes downstream EDA on `data/processed/`; it does not imply that noisy channels, missing questionnaire values, or documented recording anomalies should be removed. Any analytical exclusion must be explicit and separate from this cleaning pipeline.",
            "",
        ]
    )
    atomic_write_text(destination, "\n".join(lines))
    return verdict


def main(argv: Sequence[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw", type=Path, default=DEFAULT_RAW)
    parser.add_argument("--cleaned", type=Path, default=DEFAULT_CLEANED)
    parser.add_argument("--reports", type=Path, default=DEFAULT_REPORTS)
    parser.add_argument(
        "--rules",
        type=Path,
        default=DEFAULT_RAW.parent / "documentation_rules.json",
    )
    args = parser.parse_args(argv)
    raw, cleaned = assert_safe_paths(args.raw, args.cleaned)
    if not cleaned.is_dir():
        raise FileNotFoundError(f"Processed dataset does not exist: {cleaned}")
    if not args.rules.is_file():
        raise FileNotFoundError(f"Documentation rules do not exist: {args.rules}")

    checks, warnings, documented_exceptions = build_readiness_checks(
        raw,
        cleaned,
        args.reports,
        args.rules,
    )
    write_readiness_csv(args.reports / "eda_readiness.csv", checks)
    verdict = write_readiness_report(
        args.reports / "eda_readiness.md",
        cleaned,
        checks,
        warnings,
        documented_exceptions,
    )
    print(f"EDA readiness: {verdict}")
    return 0 if verdict == "PASS" else 2


if __name__ == "__main__":
    sys.exit(main())
